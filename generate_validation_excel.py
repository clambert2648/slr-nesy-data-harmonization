"""
generate_validation_excel.py
Génère le fichier Excel de validation pour le directeur de recherche.

Onglets :
  1. Critères         — Critères d'inclusion/exclusion (Tri 1) et grille QA (Tri 2)
  2. Inclus finaux    — 38 articles qa_pass=oui (27 corpus + 11 snowballing)
  3. Exclu Tri 2      — Articles exclus au Tri 2 (qa_pass=non, hors I4 strict)
  4. Exclu I4 strict  — Articles exclu car I4 strict échoué (175 corpus + 5 snowballing)
  5. Exclu Tri 1      — Articles exclus au Tri 1 (snowballing seulement, 30 articles)
  6. Surveys           — 59 surveys identifiés pour snowballing

Usage:
    python generate_validation_excel.py
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

# ── Paths ──
BASE = os.path.dirname(os.path.abspath(__file__))
INCLUS_PATH = os.path.join(BASE, "data", "processed", "articles_inclus.csv")
CORPUS_PATH = os.path.join(BASE, "data", "processed", "corpus_scored.csv")
SNOWBALL_PATH = os.path.join(BASE, "data", "processed", "snowballing.csv")
SURVEYS_PATH = os.path.join(BASE, "data", "processed", "articles_surveys.csv")
OUTPUT_PATH = os.path.join(BASE, "results", "validation_SLR.xlsx")

# ── SharePoint base URL for PDF links ──
SHAREPOINT_BASE = (
    "https://uqoca.sharepoint.com/sites/Projetdemmoire-NeSyHarmonisationSmantique/"
    "Documents%20partages/General/RLS/Articles"
)

# ── Build PDF location map ──
FULLTEXT_DIR = os.path.join(BASE, "data", "fulltext")
PDF_MAP = {}  # key (rank or BT01) -> subfolder/filename
for _folder in ["Extraction", "Screening_2_I4_strict", "Screening_2_excluded", "Surveys", "archive"]:
    _fpath = os.path.join(FULLTEXT_DIR, _folder)
    if os.path.isdir(_fpath):
        for _f in os.listdir(_fpath):
            if _f.lower().endswith(".pdf"):
                _key = _f.split("_")[0]  # e.g. "058" or "BT01"
                PDF_MAP[_key] = f"{_folder}/{_f}"

# ── Load data ──
inclus = pd.read_csv(INCLUS_PATH, keep_default_na=False)
corpus = pd.read_csv(CORPUS_PATH, keep_default_na=False)
snowball = pd.read_csv(SNOWBALL_PATH, keep_default_na=False)
surveys = pd.read_csv(SURVEYS_PATH, keep_default_na=False)

# ── Styles ──
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_FILL_GREEN = PatternFill(start_color="548235", end_color="548235", fill_type="solid")
HEADER_FILL_RED = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
HEADER_FILL_ORANGE = PatternFill(start_color="BF8F00", end_color="BF8F00", fill_type="solid")
HEADER_FILL_PURPLE = PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid")
HEADER_FILL_TEAL = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
TITLE_FONT = Font(bold=True, size=14, color="2F5496")
SUBTITLE_FONT = Font(bold=True, size=11, color="2F5496")
SECTION_FONT = Font(bold=True, size=12, color="548235")
WRAP_ALIGN = Alignment(wrap_text=True, vertical="top")
TOP_ALIGN = Alignment(vertical="top")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
PASS_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
FAIL_FILL = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")


def style_header(ws, row, ncols, fill=HEADER_FILL):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = fill
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        cell.border = THIN_BORDER


def style_data_rows(ws, start_row, end_row, ncols):
    for r in range(start_row, end_row + 1):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=r, column=c)
            cell.alignment = WRAP_ALIGN
            cell.border = THIN_BORDER


def auto_width(ws, max_width=60, min_width=8):
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        max_len = min_width
        for cell in col_cells[:50]:  # sample first 50 rows
            if cell.value:
                lines = str(cell.value).split("\n")
                longest = max(len(l) for l in lines)
                max_len = max(max_len, min(longest + 2, max_width))
        ws.column_dimensions[col_letter].width = max_len


def build_provenance(row, is_snowball=False, sn_row=None):
    """Build provenance string like '(R1, Scopus+IEEE)' or '(Snowballing, survey Cotovio_2023)'."""
    if is_snowball and sn_row is not None:
        stype = sn_row.get("snowball_type", "")
        src = sn_row.get("source_surveys", "")
        src_art = sn_row.get("source_articles", "")
        if stype == "survey":
            return f"Snowballing survey ({src})"
        else:
            return f"Snowballing backtracking (art. {src_art})"
    query = str(row.get("query", ""))
    db = str(row.get("database", ""))
    if query and db:
        return f"{query}, {db}"
    elif db:
        return db
    return ""


# ══════════════════════════════════════════════════════════════════════════════
# TAB 1: CRITÈRES
# ══════════════════════════════════════════════════════════════════════════════
def write_criteria_tab(wb):
    ws = wb.create_sheet("Critères", 0)

    # Title
    ws.merge_cells("A1:F1")
    ws["A1"] = "Critères de sélection — RSL Appariement neuro-symbolique"
    ws["A1"].font = TITLE_FONT

    ws.merge_cells("A2:F2")
    ws["A2"] = "Constance Lambert-Tremblay — M. Sc. informatique, UQO — Dir. Pr Étienne Gaël Tajeuna"
    ws["A2"].font = Font(italic=True, size=10)

    # ── Inclusion criteria ──
    r = 4
    ws.cell(row=r, column=1, value="CRITÈRES D'INCLUSION (Tri 1)").font = SECTION_FONT
    r += 1
    ws.cell(row=r, column=1, value="Toutes les conditions doivent être vraies, avec I2 ou I3 requis (optionnel).")
    r += 1

    inclusion = [
        ("I1", "Article de revue ou acte de conférence évalué par les pairs, 2020–2025."),
        ("I2", "Tâche d'appariement / alignement / mapping / harmonisation de données hétérogènes."),
        ("I3", "Explicabilité ou interprétabilité des décisions d'appariement (HITL). Optionnel — I2 ou I3 suffit."),
        ("I4", "Approche neuro-symbolique / hybride neural+symbolique (KG/ontologie/règles + embeddings/transformers)."),
        ("I5", "Évaluation empirique : au moins un dataset + une métrique."),
        ("I6", "Full text accessible."),
    ]
    headers_ie = ["Code", "Critère"]
    for j, h in enumerate(headers_ie, 1):
        ws.cell(row=r, column=j, value=h)
    style_header(ws, r, len(headers_ie), HEADER_FILL_GREEN)
    r += 1
    for code, desc in inclusion:
        ws.cell(row=r, column=1, value=code)
        ws.cell(row=r, column=2, value=desc)
        for c in range(1, 3):
            ws.cell(row=r, column=c).border = THIN_BORDER
            ws.cell(row=r, column=c).alignment = WRAP_ALIGN
        r += 1

    # ── Exclusion criteria ──
    r += 1
    ws.cell(row=r, column=1, value="CRITÈRES D'EXCLUSION (Tri 1)").font = SECTION_FONT
    r += 1
    ws.cell(row=r, column=1, value="Une seule condition suffit pour exclure.")
    r += 1

    exclusion = [
        ("E1", "Hors tâche : KG completion, link prediction, classification, QA, recommandation. Exception : entity alignment inter-KG si I2 explicite."),
        ("E2", "Hors modalités : images/vidéo/audio/signaux sans composante textuelle."),
        ("E3", "Méthode non algorithmique : harmonisation manuelle, discussion conceptuelle."),
        ("E4", "Pas d'évaluation : pas de résultats, pas de métriques, papier position/vision."),
        ("E5", "Type non retenu : thèse, rapport, chapitre, éditorial, poster, brevet. Surveys → snowballing."),
        ("E6", "Hors période (< 2020 ou > 2025) ou langue sans abstract exploitable."),
    ]
    for j, h in enumerate(headers_ie, 1):
        ws.cell(row=r, column=j, value=h)
    style_header(ws, r, len(headers_ie), HEADER_FILL_RED)
    r += 1
    for code, desc in exclusion:
        ws.cell(row=r, column=1, value=code)
        ws.cell(row=r, column=2, value=desc)
        for c in range(1, 3):
            ws.cell(row=r, column=c).border = THIN_BORDER
            ws.cell(row=r, column=c).alignment = WRAP_ALIGN
        r += 1

    # ── I4 Strict (Tri 2) ──
    r += 1
    ws.cell(row=r, column=1, value="I4 STRICT — Application au Tri 2").font = SECTION_FONT
    r += 1
    ws.cell(row=r, column=1, value="✅ Passe : raisonnement symbolique ACTIF couplé au neural — axiomes OWL, clauses de Horn, logique de description, logique floue formelle, règles symboliques appliquées à l'inférence.")
    ws.cell(row=r, column=1).alignment = WRAP_ALIGN
    r += 1
    ws.cell(row=r, column=1, value="❌ Échoue : KG/ontologie utilisé UNIQUEMENT comme input pour embeddings (GNN, TransE, BERT). Scores remis à 0.")
    ws.cell(row=r, column=1).alignment = WRAP_ALIGN

    # ── QA Grid (Tri 2) ──
    r += 2
    ws.cell(row=r, column=1, value="GRILLE QA — Tri 2 (seuil ≥ 3/5, Q6 bonus)").font = SECTION_FONT
    r += 1

    qa = [
        ("Q1", "Objectifs et tâche clairement définis", "0 = non | 0.5 = partiellement | 1 = oui"),
        ("Q2", "Architecture NeSy reproductible (composantes, intégration, hyperparamètres)", "0 = non | 0.5 = décrite mais pas reproductible | 1 = oui"),
        ("Q3", "Dataset nommé+taille + métrique standard (F1/P/R/MRR/Hit@k)", "0 = non | 0.5 = dataset proprio non divulgué OU métriques non standard | 1 = oui"),
        ("Q4", "Explications des décisions d'appariement (traces, scores, viz, HITL)", "0 = non | 0.5 = scores seuls, sans justification structurée | 1 = oui"),
        ("Q5", "Comparaison à baseline externe", "0 = non | 0.5 = ablation interne seulement | 1 = oui"),
        ("Q6", "[BONUS, hors total] Limites discutées", "0 = non | 0.5 = partiellement | 1 = oui"),
    ]
    qa_headers = ["Question", "Critère", "Barème"]
    for j, h in enumerate(qa_headers, 1):
        ws.cell(row=r, column=j, value=h)
    style_header(ws, r, len(qa_headers), HEADER_FILL)
    r += 1
    for code, desc, bareme in qa:
        ws.cell(row=r, column=1, value=code)
        ws.cell(row=r, column=2, value=desc)
        ws.cell(row=r, column=3, value=bareme)
        for c in range(1, 4):
            ws.cell(row=r, column=c).border = THIN_BORDER
            ws.cell(row=r, column=c).alignment = WRAP_ALIGN
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="qa_total = Q1+Q2+Q3+Q4+Q5 (max 5)").font = SUBTITLE_FONT
    r += 1
    ws.cell(row=r, column=1, value="qa_pass = \"oui\" si qa_total ≥ 3").font = SUBTITLE_FONT

    # Column widths
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 70
    ws.column_dimensions["C"].width = 55


# ══════════════════════════════════════════════════════════════════════════════
# ARTICLE SHEETS — common column builder
# ══════════════════════════════════════════════════════════════════════════════
COMMON_COLS = [
    "Rank", "Provenance", "Auteurs", "Année", "Titre", "Abstract", "Mots-clés",
    "Décision Tri 1", "Raison exclusion Tri 1",
    "Q1", "Q2", "Q3", "Q4", "Q5", "Q6", "Total QA", "QA Pass",
    "Notes QA",
    "Décision Prof.", "Rébutal", "Lien",
]

COMMON_WIDTHS = {
    "Rank": 8, "Provenance": 25, "Auteurs": 30, "Année": 6,
    "Titre": 50, "Abstract": 60, "Mots-clés": 30,
    "Décision Tri 1": 14, "Raison exclusion Tri 1": 18,
    "Q1": 5, "Q2": 5, "Q3": 5, "Q4": 5, "Q5": 5, "Q6": 5,
    "Total QA": 9, "QA Pass": 8, "Notes QA": 45,
    "Décision Prof.": 15, "Rébutal": 30, "Lien": 15,
}

LINK_FONT = Font(color="0563C1", underline="single")


def build_link(rank_str, doi="", sn_row=None):
    """Return (url, display_text) for the Lien column.
    Priority: SharePoint PDF link > DOI > empty."""
    from urllib.parse import quote

    # Normalize rank key for PDF lookup
    key = str(rank_str).replace("-", "")
    pdf_rel = PDF_MAP.get(key)
    # Try zero-padded variant for numeric ranks (e.g. "58" -> "058")
    if not pdf_rel and key.isdigit():
        pdf_rel = PDF_MAP.get(key.zfill(3))
        if not pdf_rel:
            pdf_rel = PDF_MAP.get(key.zfill(4))
    if pdf_rel:
        # SharePoint link to PDF
        parts = pdf_rel.split("/", 1)
        folder_enc = quote(parts[0], safe="")
        file_enc = quote(parts[1], safe="")
        url = f"{SHAREPOINT_BASE}/{folder_enc}/{file_enc}"
        return url, "PDF"

    # Fallback: DOI link
    if sn_row is not None:
        doi = str(sn_row.get("doi", doi)).strip()
    doi = str(doi).strip()
    if doi:
        if doi.startswith("http"):
            return doi, "Web"
        return f"https://doi.org/{doi}", "DOI"

    return "", ""


def article_to_row(row, provenance="", sn_row=None):
    """Convert a DataFrame row to the common column values."""
    # Determine QA notes (join all q notes)
    notes_parts = []
    for q in ["qa_notes_q1", "qa_notes_q2", "qa_notes_q3", "qa_notes_q4", "qa_notes_q5", "qa_notes_q6"]:
        val = str(row.get(q, "")).strip()
        if val:
            notes_parts.append(val)
    qa_notes = " | ".join(notes_parts) if notes_parts else ""

    # For snowballing articles, map field names
    if sn_row is not None:
        rank = str(sn_row.get("id", row.get("rank", "")))
        decision_tri1 = str(sn_row.get("tri1", ""))
        excl_reason = ""
        if "exclu_" in decision_tri1:
            excl_reason = decision_tri1.replace("exclu_", "")
            decision_tri1 = "exclu"
        qa_notes_sn = str(sn_row.get("notes", ""))
        if qa_notes_sn and not qa_notes:
            qa_notes = qa_notes_sn
    else:
        rank = str(row.get("rank", ""))
        decision_tri1 = str(row.get("decision", ""))
        excl_reason = str(row.get("exclusion_reason", ""))

    # Build link
    doi_val = str(row.get("doi", "")).strip()
    link_url, link_text = build_link(rank, doi_val, sn_row)

    return [
        rank,
        provenance,
        str(row.get("authors", ""))[:200],
        row.get("year", ""),
        str(row.get("title", "")),
        str(row.get("abstract", "")),
        str(row.get("keywords", "")),
        decision_tri1,
        excl_reason,
        row.get("qa_q1", ""),
        row.get("qa_q2", ""),
        row.get("qa_q3", ""),
        row.get("qa_q4", ""),
        row.get("qa_q5", ""),
        row.get("qa_q6", ""),
        row.get("qa_total", ""),
        row.get("qa_pass", ""),
        qa_notes[:500],
        "",  # Décision Prof.
        "",  # Rébutal
        (link_url, link_text),  # Lien — tuple for hyperlink
    ]


def write_article_sheet(wb, sheet_name, rows_data, header_fill, explanation=None):
    """Write an article sheet with common columns."""
    ws = wb.create_sheet(sheet_name)
    start_row = 1

    if explanation:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COMMON_COLS))
        ws["A1"] = explanation
        ws["A1"].font = Font(italic=True, size=10, color="C00000")
        ws["A1"].alignment = WRAP_ALIGN
        ws.row_dimensions[1].height = 40
        start_row = 3

    # Header
    for j, col_name in enumerate(COMMON_COLS, 1):
        ws.cell(row=start_row, column=j, value=col_name)
    style_header(ws, start_row, len(COMMON_COLS), header_fill)

    # Data
    link_col = len(COMMON_COLS)  # last column = Lien
    for i, row_data in enumerate(rows_data, start_row + 1):
        for j, val in enumerate(row_data, 1):
            cell = ws.cell(row=i, column=j)
            if j == link_col and isinstance(val, tuple) and len(val) == 2:
                url, display = val
                if url:
                    cell.value = display
                    cell.hyperlink = url
                    cell.font = LINK_FONT
                else:
                    cell.value = ""
            else:
                cell.value = val

    # Style data
    if rows_data:
        style_data_rows(ws, start_row + 1, start_row + len(rows_data), len(COMMON_COLS))

    # Conditional row coloring for pass/fail
    for i, row_data in enumerate(rows_data, start_row + 1):
        qa_pass_val = row_data[16] if len(row_data) > 16 else ""
        fill = None
        if qa_pass_val == "oui":
            fill = PASS_FILL
        elif qa_pass_val == "non":
            fill = FAIL_FILL
        if fill:
            for c in range(1, len(COMMON_COLS) + 1):
                cell = ws.cell(row=i, column=c)
                cell.fill = fill
                # Preserve hyperlink font on Lien column
                if c == link_col and cell.hyperlink:
                    cell.font = LINK_FONT

    # Column widths
    for j, col_name in enumerate(COMMON_COLS, 1):
        ws.column_dimensions[get_column_letter(j)].width = COMMON_WIDTHS.get(col_name, 12)

    # Freeze panes
    ws.freeze_panes = ws.cell(row=start_row + 1, column=1)

    # Auto-filter
    ws.auto_filter.ref = f"A{start_row}:{get_column_letter(len(COMMON_COLS))}{start_row + len(rows_data)}"

    return ws


# ══════════════════════════════════════════════════════════════════════════════
# BUILD DATA FOR EACH TAB
# ══════════════════════════════════════════════════════════════════════════════

# Build snowball lookup
sn_lookup = {}
for _, sn_r in snowball.iterrows():
    sn_lookup[str(sn_r["id"]).replace("-", "")] = sn_r

# ── TAB 2: Inclus finaux (38 articles, qa_pass=oui) ──
inclus_final = inclus[inclus["qa_pass"] == "oui"].copy()
inclus_final_rows = []
for _, row in inclus_final.iterrows():
    rank_str = str(row["rank"])
    sn_row = sn_lookup.get(rank_str) if not rank_str.isdigit() else None
    if sn_row is not None:
        prov = build_provenance(row, is_snowball=True, sn_row=sn_row)
    else:
        prov = build_provenance(row)
    inclus_final_rows.append(article_to_row(row, prov, sn_row))

# ── TAB 3: Exclu Tri 2 (qa_pass=non, NOT I4 strict) — corpus articles ──
non_pass = inclus[inclus["qa_pass"] == "non"].copy()
exclu_tri2_real = non_pass[~non_pass["qa_notes_q2"].str.contains("I4", case=False, na=False)]
# Also add snowballing exclu_tri2 that are not I4
sn_exclu_tri2 = snowball[(snowball["tri2"].str.startswith("exclu")) & (~snowball["tri2"].str.contains("I4", case=False))]

exclu_tri2_rows = []
for _, row in exclu_tri2_real.iterrows():
    rank_str = str(row["rank"])
    sn_row = sn_lookup.get(rank_str) if not rank_str.isdigit() else None
    prov = build_provenance(row, is_snowball=(sn_row is not None), sn_row=sn_row)
    exclu_tri2_rows.append(article_to_row(row, prov, sn_row))

for _, sn_r in sn_exclu_tri2.iterrows():
    sid = str(sn_r["id"]).replace("-", "")
    # Build a minimal row dict from snowball data
    row_dict = {
        "rank": sn_r["id"], "authors": sn_r.get("authors", ""), "year": sn_r.get("year", ""),
        "title": sn_r.get("title", ""), "abstract": sn_r.get("abstract", ""),
        "keywords": "", "doi": sn_r.get("doi", ""), "decision": "", "exclusion_reason": "",
        "qa_q1": "", "qa_q2": "", "qa_q3": "", "qa_q4": "", "qa_q5": "", "qa_q6": "",
        "qa_total": "", "qa_pass": "", "qa_notes_q1": "", "qa_notes_q2": sn_r.get("notes", ""),
    }
    prov = build_provenance(row_dict, is_snowball=True, sn_row=sn_r)
    exclu_tri2_rows.append(article_to_row(pd.Series(row_dict), prov, sn_r))

# ── TAB 4: Exclu I4 strict (175 corpus + 5 snowballing) ──
exclu_i4_corpus = non_pass[non_pass["qa_notes_q2"].str.contains("I4", case=False, na=False)]
sn_exclu_i4 = snowball[snowball["tri2"].str.contains("I4", case=False, na=False)]

exclu_i4_rows = []
for _, row in exclu_i4_corpus.iterrows():
    rank_str = str(row["rank"])
    sn_row = sn_lookup.get(rank_str) if not rank_str.isdigit() else None
    prov = build_provenance(row, is_snowball=(sn_row is not None), sn_row=sn_row)
    exclu_i4_rows.append(article_to_row(row, prov, sn_row))

for _, sn_r in sn_exclu_i4.iterrows():
    sid = str(sn_r["id"]).replace("-", "")
    # Check if already in articles_inclus (to avoid duplicates)
    if sid in inclus["rank"].astype(str).values:
        continue
    row_dict = {
        "rank": sn_r["id"], "authors": sn_r.get("authors", ""), "year": sn_r.get("year", ""),
        "title": sn_r.get("title", ""), "abstract": sn_r.get("abstract", ""),
        "keywords": "", "doi": sn_r.get("doi", ""), "decision": "", "exclusion_reason": "",
        "qa_q1": "", "qa_q2": "", "qa_q3": "", "qa_q4": "", "qa_q5": "", "qa_q6": "",
        "qa_total": "", "qa_pass": "", "qa_notes_q1": "", "qa_notes_q2": sn_r.get("notes", ""),
    }
    prov = build_provenance(row_dict, is_snowball=True, sn_row=sn_r)
    exclu_i4_rows.append(article_to_row(pd.Series(row_dict), prov, sn_r))

# ── TAB 5: Exclu Tri 1 (corpus + snowballing) ──
# 5a) Corpus principal — decision=exclude
corpus_exclu_tri1 = corpus[corpus["decision"] == "exclude"].copy()
exclu_tri1_rows = []
for _, row in corpus_exclu_tri1.iterrows():
    prov = build_provenance(row)
    exclu_tri1_rows.append(article_to_row(row, prov))

# 5b) Snowballing — tri1 starts with "exclu"
sn_exclu_tri1 = snowball[snowball["tri1"].str.startswith("exclu")]
for _, sn_r in sn_exclu_tri1.iterrows():
    row_dict = {
        "rank": sn_r["id"], "authors": sn_r.get("authors", ""), "year": sn_r.get("year", ""),
        "title": sn_r.get("title", ""), "abstract": sn_r.get("abstract", ""),
        "keywords": "", "doi": sn_r.get("doi", ""), "decision": "", "exclusion_reason": "",
        "qa_q1": "", "qa_q2": "", "qa_q3": "", "qa_q4": "", "qa_q5": "", "qa_q6": "",
        "qa_total": "", "qa_pass": "", "qa_notes_q1": "", "qa_notes_q2": sn_r.get("notes", ""),
    }
    prov = build_provenance(row_dict, is_snowball=True, sn_row=sn_r)
    exclu_tri1_rows.append(article_to_row(pd.Series(row_dict), prov, sn_r))

# ── TAB 6: Surveys (59 surveys from corpus) ──
survey_rows = []
for _, row in surveys.iterrows():
    rank_val = row.get("rank", row.get("Rank", row.get("Unnamed: 0", "")))
    prov_row = None
    # Get provenance from corpus_scored
    if rank_val != "":
        matches = corpus[corpus["rank"] == float(rank_val)] if str(rank_val).replace(".", "").isdigit() else pd.DataFrame()
        if len(matches) > 0:
            prov_row = matches.iloc[0]
    prov = build_provenance(prov_row if prov_row is not None else row, is_snowball=False)
    row_dict = {
        "rank": rank_val, "authors": row.get("authors", ""), "year": row.get("year", ""),
        "title": row.get("title", ""), "abstract": row.get("abstract", ""),
        "keywords": row.get("keywords", ""), "doi": row.get("doi", ""),
        "decision": "survey", "exclusion_reason": "E5s",
        "database": row.get("database", prov_row["database"] if prov_row is not None else ""),
        "query": row.get("query", prov_row["query"] if prov_row is not None else ""),
        "qa_q1": "", "qa_q2": "", "qa_q3": "", "qa_q4": "", "qa_q5": "", "qa_q6": "",
        "qa_total": "", "qa_pass": "", "qa_notes_q1": "", "qa_notes_q2": "",
    }
    survey_rows.append(article_to_row(pd.Series(row_dict), prov))


# ══════════════════════════════════════════════════════════════════════════════
# WRITE WORKBOOK
# ══════════════════════════════════════════════════════════════════════════════
wb = Workbook()
# Remove default sheet
wb.remove(wb.active)

write_criteria_tab(wb)

write_article_sheet(
    wb, f"Inclus finaux ({len(inclus_final_rows)})",
    inclus_final_rows, HEADER_FILL_GREEN,
)

write_article_sheet(
    wb, f"Exclu Tri 2 ({len(exclu_tri2_rows)})",
    exclu_tri2_rows, HEADER_FILL_RED,
    explanation="Articles inclus au Tri 1, mais exclus au Tri 2 (score QA < 3/5 ou critère non satisfait). Exclut les I4 strict (onglet séparé).",
)

write_article_sheet(
    wb, f"Exclu I4 strict ({len(exclu_i4_rows)})",
    exclu_i4_rows, HEADER_FILL_ORANGE,
    explanation="Articles inclus au Tri 1 sur titre/abstract, mais exclu au Tri 2 car le critère I4 strict échoue à la lecture du full text : "
                "le KG/ontologie est utilisé UNIQUEMENT comme input pour embeddings (GNN, TransE, BERT), sans raisonnement symbolique actif. "
                "Tous les scores QA sont remis à 0.",
)

write_article_sheet(
    wb, f"Exclu Tri 1 ({len(exclu_tri1_rows)})",
    exclu_tri1_rows, HEADER_FILL_PURPLE,
    explanation="Articles exclus au Tri 1 (titre/abstract) : corpus principal (decision=exclude) + candidats snowballing exclus. "
                "Raisons : E1 (hors tâche), E2 (hors modalités), E3 (non algorithmique), E4 (pas d'évaluation), E5 (type non retenu), E6 (hors période).",
)

write_article_sheet(
    wb, f"Surveys ({len(survey_rows)})",
    survey_rows, HEADER_FILL_TEAL,
    explanation="Surveys identifiés dans le corpus principal. Utilisés pour le snowballing (identification de candidats supplémentaires).",
)

# Save
wb.save(OUTPUT_PATH)
print(f"Excel généré : {OUTPUT_PATH}")
print(f"  Inclus finaux : {len(inclus_final_rows)}")
print(f"  Exclu Tri 2 :   {len(exclu_tri2_rows)}")
print(f"  Exclu I4 strict: {len(exclu_i4_rows)}")
print(f"  Exclu Tri 1 SB : {len(exclu_tri1_rows)}")
print(f"  Surveys :        {len(survey_rows)}")
